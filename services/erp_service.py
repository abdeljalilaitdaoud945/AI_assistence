"""
Service ERP — lecture des données réelles depuis erp_data.xlsx.

Le fichier erp_data.xlsx (à la racine du projet) contient 5 onglets :
  - Finance              : KPI (CA, marge, trésorerie, taux d'occupation)
  - FacturesImpayees     : détail des factures, calcul total via formule SUMIF
  - Projets              : projets en retard avec cause et impact
  - Ventes               : CA par service avec croissance
  - README               : instructions pour l'utilisateur

L'utilisateur peut modifier le fichier dans Excel : l'app le relit à chaque
ouverture de la vue ERP. Si le fichier est absent ou corrompu, fallback
automatique sur les données simulées (MOCK_DATA) pour éviter tout crash.
"""

import json
from pathlib import Path
from google import genai


# =====================================================================
# Localisation du fichier de données
# =====================================================================
def _find_data_file() -> Path:
    here = Path(__file__).resolve().parent
    candidates = [
        here / "erp_data.xlsx",
        here.parent / "erp_data.xlsx",
        Path.cwd() / "erp_data.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return candidates[1]


DATA_FILE = _find_data_file()


# =====================================================================
# Fallback simulé (utilisé si erp_data.xlsx absent)
# =====================================================================
MOCK_DATA = {
    "finance": {
        "chiffre_affaires": {"realise_ytd": 12500000, "objectif_ytd": 15000000,
                             "tendance_mensuelle": "baisse (-5%)", "devise": "MAD"},
        "marge_globale": {"realise": 22.5, "objectif": 28.0, "unite": "%"},
        "tresorerie": {"disponible": 850000, "seuil_alerte": 1000000,
                       "statut": "critique"},
        "factures_impayees": {"montant_total": 420000,
                              "clients_principaux": ["Groupe ONA",
                                                     "Ministère de la Santé",
                                                     "TechCorp"]},
    },
    "operations": {
        "projets_en_retard": [
            {"client": "TechCorp", "projet": "Migration Cloud AWS",
             "retard_jours": 18, "impact_estime": 120000,
             "cause": "Manque de ressources DevOps seniors"},
            {"client": "MegaBank", "projet": "Audit de Sécurité SI",
             "retard_jours": 7, "impact_estime": 45000,
             "cause": "Attente des accès VPN côté client"},
            {"client": "AutoMaroc", "projet": "Déploiement ERP",
             "retard_jours": 25, "impact_estime": 250000,
             "cause": "Changement de périmètre (Scope creep)"},
        ],
        "top_ventes_services": [
            {"service": "Intégration Cybersécurité", "ca_genere": 4500000,
             "croissance": "+15%"},
            {"service": "Consulting Cloud", "ca_genere": 3200000,
             "croissance": "+8%"},
            {"service": "Support & Maintenance", "ca_genere": 1800000,
             "croissance": "-2%"},
        ],
        "taux_occupation_equipes": 92.5,
    },
    "_source": "MOCK (fichier erp_data.xlsx introuvable)",
}


def _load_from_xlsx(path: Path) -> dict:
    """Charge les données depuis erp_data.xlsx. Lève une exception si invalide."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)

    # --- Finance ---
    ws = wb["Finance"]
    fin_rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        label, real, obj, unit = row[0], row[1], row[2], row[3]
        fin_rows[label] = {"realise": real, "objectif": obj, "unite": unit}

    # Recherche tolérante : on normalise les libellés pour matcher quoi qu'il
    # arrive (apostrophe courbe vs droite, espaces multiples, casse, accents)
    import unicodedata
    def _norm(s):
        if not s:
            return ""
        s = str(s).lower().strip()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        # Remplace apostrophes typographiques par simple
        s = s.replace("'", "'").replace("’", "'")
        # Espaces multiples → un seul
        return " ".join(s.split())

    def _find(needle):
        n = _norm(needle)
        for k, v in fin_rows.items():
            if n in _norm(k):
                return v
        return {}

    ca = _find("chiffre d'affaires")
    marge = _find("marge")
    treso = _find("tresorerie")
    occup = _find("occupation")

    # Diagnostic console (utile si tout est à 0)
    if not ca:
        print(f"[erp_service] ⚠️ Libellé 'Chiffre d'affaires' introuvable. "
              f"Libellés détectés : {list(fin_rows.keys())}")

    treso_real = float(treso.get("realise") or 0)
    treso_seuil = float(treso.get("objectif") or 1)
    treso_statut = ("critique" if treso_real < treso_seuil * 0.85
                    else "tendu" if treso_real < treso_seuil
                    else "ok")

    # --- Factures ---
    ws = wb["FacturesImpayees"]
    factures_list = []
    total_impayes = 0.0
    clients_set = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or str(row[0]).startswith("TOTAL"):
            continue
        if len(row) < 7:
            continue
        statut = row[6]
        if statut == "En retard":
            montant = float(row[4] or 0)
            total_impayes += montant
            client = row[1]
            if client and client not in clients_set:
                clients_set.append(client)
            factures_list.append({
                "numero": row[0], "client": row[1],
                "montant": montant, "retard_jours": row[5],
            })

    # --- Projets ---
    ws = wb["Projets"]
    projets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or str(row[0]).startswith("IMPACT"):
            continue
        if len(row) < 6:
            continue
        projets.append({
            "client": row[1],
            "projet": row[2],
            "retard_jours": int(row[3] or 0),
            "impact_estime": int(row[4] or 0),
            "cause": row[5],
        })

    # --- Ventes ---
    ws = wb["Ventes"]
    ventes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or str(row[0]).startswith("TOTAL"):
            continue
        if len(row) < 3:
            continue
        croiss = row[2]
        try:
            cv = float(croiss)
            croiss_str = f"{'+' if cv >= 0 else ''}{cv:.1f}%"
        except Exception:
            croiss_str = str(croiss) if croiss else "0%"
        ventes.append({
            "service": row[0],
            "ca_genere": int(row[1] or 0),
            "croissance": croiss_str,
        })

    return {
        "finance": {
            "chiffre_affaires": {
                "realise_ytd": int(ca.get("realise") or 0),
                "objectif_ytd": int(ca.get("objectif") or 0),
                "tendance_mensuelle": "—",
                "devise": "MAD",
            },
            "marge_globale": {
                "realise": float(marge.get("realise") or 0),
                "objectif": float(marge.get("objectif") or 0),
                "unite": "%",
            },
            "tresorerie": {
                "disponible": int(treso_real),
                "seuil_alerte": int(treso_seuil),
                "statut": treso_statut,
            },
            "factures_impayees": {
                "montant_total": int(total_impayes),
                "clients_principaux": clients_set[:5],
                "detail": factures_list,
            },
        },
        "operations": {
            "projets_en_retard": projets,
            "top_ventes_services": ventes,
            "taux_occupation_equipes": float(occup.get("realise") or 0),
        },
        "_source": f"LIVE ({path.name})",
    }


def get_erp_data() -> dict:
    """Retourne les données ERP. Lit le .xlsx si présent, sinon fallback mock."""
    if DATA_FILE.exists():
        try:
            return _load_from_xlsx(DATA_FILE)
        except Exception as e:
            print(f"[erp_service] Erreur lecture {DATA_FILE.name}: {e} "
                  f"-> fallback simulé")
    return MOCK_DATA


def get_data_source() -> str:
    return get_erp_data().get("_source", "?")


def get_erp_summary_text() -> str:
    data = get_erp_data()
    fin = data["finance"]
    ops = data["operations"]
    ca = fin["chiffre_affaires"]

    lines = [f"📊 Synthèse ERP — source : {data.get('_source', '?')}\n"]
    lines.append(f"• CA YTD : {ca['realise_ytd']:,} / {ca['objectif_ytd']:,} {ca['devise']}")
    lines.append(f"• Marge : {fin['marge_globale']['realise']}% "
                 f"(Obj: {fin['marge_globale']['objectif']}%)")
    lines.append(f"• Trésorerie : {fin['tresorerie']['disponible']:,} {ca['devise']} "
                 f"(Statut: {fin['tresorerie']['statut']})")
    lines.append(f"• Factures impayées : {fin['factures_impayees']['montant_total']:,} "
                 f"{ca['devise']}")

    retards = ops.get("projets_en_retard", [])
    if retards:
        lines.append("\n⚠️ Projets critiques en retard :")
        for r in retards:
            lines.append(f"  - {r['projet']} ({r['client']}) : {r['retard_jours']}j | "
                         f"Cause: {r['cause']} | "
                         f"Impact: {r['impact_estime']:,} {ca['devise']}")
    return "\n".join(lines)


def generate_erp_recommendations() -> str:
    try:
        client = genai.Client()
        data = {k: v for k, v in get_erp_data().items() if k != "_source"}
        data_str = json.dumps(data, indent=2, ensure_ascii=False)
        prompt = (
            "Tu es le Directeur Général Adjoint (stratégie et finance) d'une "
            "société de conseil/IT marocaine. Analyse ces données ERP réelles :\n"
            f"{data_str}\n\n"
            "Rédige une note stratégique percutante (max 280 mots) avec EXACTEMENT "
            "ce format, sans introduction ni conclusion :\n\n"
            "🔴 ALERTE FINANCIÈRE\n"
            "- (3 points max, bullets courts, factuels, chiffrés)\n\n"
            "🟠 RISQUES OPÉRATIONNELS\n"
            "- (3 points max, mentionne projets/clients/causes)\n\n"
            "🟢 PLAN D'ACTION IMMÉDIAT\n"
            "- (3 directives actionnables court terme)\n\n"
            "Règles strictes : chaque section commence EXACTEMENT par son emoji et son titre. "
            "Chaque bullet commence par '- '. Pas de texte hors sections. "
            "Vocabulaire pilotage CODIR, ton ferme et chiffré."
        )
        response = client.models.generate_content(
            model="gemini-2.5-flash", contents=prompt)
        return response.text
    except Exception as e:
        return f"Erreur lors de l'analyse IA : {e}"